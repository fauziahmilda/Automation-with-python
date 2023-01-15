import openpyxl as xl  # import the library
from openpyxl.chart import BarChart, Reference


# add function to create ability upload any file, not rely on one specify file


def process_workbook(filename):
    wb = xl.load_workbook(filename)  # get the workbook
    sheet = wb['Sheet1']  # get the sheet
    cell = sheet['a1']  # get the cell
    cell2 = sheet.cell(1, 1)  # another way to get cell
    print(cell.value)

    print(sheet.max_row)  # get the max row value

    for row in range(2, sheet.max_row + 1):  # include 2, 3, 4
        cell3 = sheet.cell(row, 3)  # cell in price section
        print(cell3.value)
        corrected_price = cell3.value * 0.9
        # add to new collumn
        corrected_price_cell = sheet.cell(row, 4)  # get the collumn, new place
        corrected_price_cell.value = corrected_price  # put the value into the column

    values = Reference(sheet,
                       min_row=2,  # select the cell
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()  # create the chart
    chart.add_data(values)  # put the value on the chart
    sheet.add_chart(chart, 'e2')  # put the chart on the sheet

    wb.save(filename)  # then save the new update here
